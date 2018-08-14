from copy import copy
from typing import NamedTuple

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.styles import Font, Alignment
from openpyxl.styles.cell_style import CellStyle
from openpyxl.styles.colors import RED
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter

from app.constants import Files


# class TokensUpdater:
#
#     def __init__(self, tokens_file, answers_file):
#         pass
#
#     def update_session(self, session: int):
#         pass
#
#     def update_sessions(self, sessions: []):
#         for session in sessions:
#             self.update_session(session)


class TokensUpdater:
    nicknames_column = column_index_from_string(Files.TOKENS_NICKNAMES_COL)

    def __init__(self, file):
        self.file = file
        self.excel = load_workbook(file, data_only=False)
        self.sheet = self.excel[Files.TOKENS_SHEET]

    def update_user_tokens(self, nickname_row, session_column, tokens_value):
        tokens_cell = self.sheet.cell(row=nickname_row, column=session_column)
        tokens_cell.value = tokens_value
        tokens_cell.alignment = Alignment(horizontal='center')

    def update_users_formula(self):
        formula_column = column_index_from_string(Files.TOKENS_CURRENT_COL)
        tokens_first_column = get_column_letter(self.nicknames_column + 1)
        last_column = get_column_letter(self.sheet.max_column)

        for row in self.sheet.iter_rows(min_row=2, min_col=formula_column, max_col=formula_column):
            current_cell = row[0]
            formula = current_cell.value
            if formula:
                tok = Tokenizer(formula)
                tok.items[1].value = '{}{}:{}{}'.format(tokens_first_column, current_cell.row, last_column,
                                                        current_cell.row)
                current_cell.value = tok.render()

    def update_user_bonuses(self):
        pass

    def create_nickname_row(self, nickname):
        last_nickname_row = self.sheet.max_row
        new_nickname_row = last_nickname_row + 1
        self.sheet.insert_rows(new_nickname_row)
        self.sheet.cell(row=new_nickname_row,
                        column=self.nicknames_column).value = nickname
        self.create_formula_for_nickname(new_nickname_row)
        CellCopy(self.sheet).copy_nickname_row_style_from_above_row(new_nickname_row)
        self.save_file()

        return new_nickname_row

    def create_formula_for_nickname(self, nickname_row):
        tokens_first_column = get_column_letter(self.nicknames_column + 1)
        last_column = get_column_letter(self.sheet.max_column)
        current_tokens_column = column_index_from_string(Files.TOKENS_CURRENT_COL)
        bonus_tokens_column = Files.TOKENS_BONUS_COL
        used_tokens_column = Files.TOKENS_USED_COL
        self.sheet.cell(row=nickname_row, column=current_tokens_column).value = '=SUM({}{}:{}{})+{}{}-{}{}'.format(
            tokens_first_column, nickname_row, last_column, nickname_row, bonus_tokens_column,
            nickname_row, used_tokens_column, nickname_row)

    def find_nickname_row(self, nickname):
        nickname_row = None
        nickname = nickname.strip()
        for row in self.sheet.iter_rows(min_row=2, min_col=self.nicknames_column, max_col=self.nicknames_column):
            current_cell = row[0]
            current_nickname = current_cell.value
            if current_nickname:
                current_nickname = current_nickname.strip()
                if (current_nickname == nickname):
                    nickname_row = current_cell.row
                    break

        return nickname_row

    def find_session_column(self, id):
        """Find column where to place current session."""
        # check if it exists already
        # place it if it doesn't
        tokens_first_column = self.nicknames_column + 1
        session_column = None
        new_column_placement = tokens_first_column

        for col in self.sheet.iter_cols(min_col=tokens_first_column, max_row=1):
            column_session_number = int(col[0].value.replace('#', ''))
            current_column_index = column_index_from_string(col[0].column)

            if id < column_session_number:
                new_column_placement = current_column_index + 1

            if column_session_number == int(id):
                session_column = current_column_index
                break

        return session_column, new_column_placement

    def create_session_column(self, column, session_id):
        self.sheet.insert_cols(column)
        session_header_cell = self.sheet.cell(1, column)
        session_header_cell.value = '#' + str(session_id)
        CellCopy(self.sheet).copy_session_header_from_previous(session_header_cell).add_session_borders(
            session_header_cell).fix_last_column_style()

        self.save_file()

    def save_file(self):
        self.excel.save(self.file)


class CellCopy:

    def __init__(self, sheet):
        self.sheet = sheet

    def fix_last_column_style(self):
        self.sheet.column_dimensions[get_column_letter(self.sheet.max_column)].width = 4
        return self

    def copy_session_header_from_previous(self, session_header_cell):
        properties = ['font', 'fill', 'border', 'alignment', 'number_format', 'protection']
        previous_session_column = column_index_from_string(session_header_cell.column) + 1
        previous_session_header_cell = self.sheet.cell(1, previous_session_column)

        self.__copy_properties(previous_session_header_cell, session_header_cell, properties)

        return self

    def copy_nickname_row_style_from_above_row(self, nickname_row):
        properties = ['font', 'border', 'alignment', 'number_format', 'protection']

        for row in self.sheet.iter_rows(min_row=nickname_row, max_row=nickname_row):
            for cell in row:
                cell_above = self.sheet.cell(cell.row - 1, column_index_from_string(cell.column))
                self.__copy_properties(cell_above, cell, properties)

    def add_session_borders(self, session_header_cell):
        properties = ['border', 'alignment', 'number_format', 'protection']
        session_column = column_index_from_string(session_header_cell.column)
        previous_session_row = session_header_cell.row + 1
        previous_session_column = column_index_from_string(session_header_cell.column) + 1
        previous_session_first_result_cell = self.sheet.cell(previous_session_row, previous_session_column)

        for row in self.sheet.iter_rows(min_row=2, min_col=session_column, max_col=session_column):
            for cell in row:
                self.__copy_properties(previous_session_first_result_cell, cell, properties)

        return self

    def __copy_properties(self, from_cell, to_cell, properties):
        for cell_property in properties:
            setattr(to_cell, cell_property, copy(getattr(from_cell, cell_property)))

        return self


class AnswersParser:
    def __init__(self, file: str, sessions_to_extract: list):
        self.sessions_to_extract = sessions_to_extract
        self.excel = load_workbook(file, data_only=True)

    def __extract_session_data(self, number):
        sheet = self.excel['#' + str(number)]
        session_data = []
        tokens_column = column_index_from_string(Files.ANSWERS_TOKENS_COL) - 1
        nicknames_column = column_index_from_string(Files.ANSWERS_NICKNAMES_COL) - 1

        for row in sheet.iter_rows(min_row=2):
            # todo throw error or skip - tokens column not defined
            if row[tokens_column].value:
                session_data.append((row[nicknames_column].value, row[tokens_column].value))

        return session_data

    def extract_answers_data(self):
        answers_data = {}

        for session_no in self.sessions_to_extract:
            answers_data[session_no] = self.__extract_session_data(session_no)

        return answers_data


class Result(NamedTuple):
    nickname: int
    answer: str


class Results(NamedTuple):
    pass


class SessionResults(NamedTuple):
    session: int
    results: Result
